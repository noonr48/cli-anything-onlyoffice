import csv
import os
import tempfile
import unittest
from unittest import mock

from openpyxl import Workbook, load_workbook

from cli_anything.onlyoffice.utils.docserver import get_client


class OnlyOfficeXLSXOpsTests(unittest.TestCase):
    def setUp(self):
        self.client = get_client()
        self.ops = self.client._xlsx_ops
        self.tmpdir = tempfile.TemporaryDirectory(prefix="oo-xlsx-ops-test-")
        self.base = self.tmpdir.name

    def tearDown(self):
        self.tmpdir.cleanup()

    def _path(self, name: str) -> str:
        return os.path.join(self.base, name)

    def test_xlsx_ops_create_write_and_sheet_list_roundtrip(self):
        path = self._path("roundtrip.xlsx")

        create_result = self.ops.create_spreadsheet(path, sheet_name="Data")
        write_result = self.ops.write_spreadsheet(
            path,
            ["Name", "Score"],
            [["Alice", "90"], ["Bob", "85"]],
            sheet_name="Data",
        )
        list_result = self.ops.sheet_list(path)

        self.assertTrue(create_result["success"])
        self.assertTrue(write_result["success"])
        self.assertTrue(list_result["success"])
        self.assertEqual(list_result["sheet_count"], 1)
        self.assertEqual(list_result["sheets"][0]["name"], "Data")

        wb = load_workbook(path)
        ws = wb["Data"]
        self.assertEqual(ws["A2"].value, "Alice")
        self.assertEqual(ws["B2"].value, 90)
        self.assertEqual(ws["A3"].value, "Bob")
        self.assertEqual(ws["B3"].value, 85)
        wb.close()

    def test_xlsx_ops_formula_audit_detects_unsupported_functions(self):
        path = self._path("formula_audit.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit"
        ws["A1"] = 10
        ws["A2"] = 20
        ws["B1"] = "=SUM(A1:A2)"
        ws["B2"] = "=XLOOKUP(A1,A1:A2,A1:A2)"
        wb.save(path)
        wb.close()

        result = self.ops.audit_spreadsheet_formulas(path, sheet_name="Audit")

        self.assertTrue(result["success"])
        self.assertEqual(result["formula_count"], 2)
        self.assertIn("XLOOKUP", result["unsupported_functions"])
        self.assertFalse(result["safe_for_cli_formula_eval"])
        self.assertGreaterEqual(len(result["risks"]), 1)

    def test_xlsx_ops_write_neutralizes_formula_like_text_and_resolves_text_columns(self):
        path = self._path("neutralized_write.xlsx")

        result = self.ops.write_spreadsheet(
            path,
            ["ID", "Value", "Comment"],
            [["00123", "=1+1", "+cmd"], ["00045", "42", "-cmd"]],
            sheet_name="Data",
            overwrite_workbook=True,
            text_columns=["ID", "C"],
        )

        self.assertTrue(result["success"])
        self.assertEqual(result["text_columns"], [1, 3])
        self.assertEqual(result["neutralized_cells"], 3)

        wb = load_workbook(path, data_only=False)
        ws = wb["Data"]
        self.assertEqual(ws["A2"].value, "00123")
        self.assertEqual(ws["B2"].value, "'=1+1")
        self.assertEqual(ws["B2"].data_type, "s")
        self.assertEqual(ws["B3"].value, 42)
        self.assertEqual(ws["C2"].value, "'+cmd")
        self.assertEqual(ws["C3"].value, "'-cmd")
        wb.close()

        audit = self.ops.audit_spreadsheet_formulas(path, sheet_name="Data")
        self.assertTrue(audit["success"])
        self.assertEqual(audit["formula_count"], 0)

        formula_result = self.ops.add_formula(path, "B4", "=SUM(B3:B3)", sheet_name="Data")
        self.assertTrue(formula_result["success"])
        audit = self.ops.audit_spreadsheet_formulas(path, sheet_name="Data")
        self.assertTrue(audit["success"])
        self.assertEqual(audit["formula_count"], 1)

    def test_xlsx_ops_append_and_cell_write_neutralize_formula_like_text(self):
        path = self._path("neutralized_cells.xlsx")
        self.ops.create_spreadsheet(path, sheet_name="Data")

        append = self.ops.append_to_spreadsheet(
            path, ["=1+1", "@cmd", "-cmd"], sheet_name="Data"
        )
        cell = self.ops.cell_write(path, "D1", "=2+2", sheet_name="Data")

        self.assertTrue(append["success"])
        self.assertEqual(append["neutralized_cells"], 3)
        self.assertTrue(cell["success"])
        self.assertTrue(cell["neutralized"])

        wb = load_workbook(path, data_only=False)
        ws = wb["Data"]
        self.assertEqual(ws["A2"].value, "'=1+1")
        self.assertEqual(ws["B2"].value, "'@cmd")
        self.assertEqual(ws["C2"].value, "'-cmd")
        self.assertEqual(ws["D1"].value, "'=2+2")
        self.assertEqual(ws["D1"].data_type, "s")
        wb.close()

    def test_xlsx_ops_csv_import_export_neutralizes_and_export_uses_locks(self):
        path = self._path("csv_safe.xlsx")
        csv_input = self._path("unsafe.csv")
        csv_output = self._path("safe.csv")
        with open(csv_input, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["ID", "FormulaLike", "AtSign"])
            writer.writerow(["00123", "=1+1", "@cmd"])
            writer.writerow(["00045", "+cmd", "-cmd"])

        imported = self.ops.csv_import(
            path, csv_input, sheet_name="Data", text_columns=["A"]
        )
        self.assertTrue(imported["success"])
        self.assertEqual(imported["text_columns"], [1])
        self.assertEqual(imported["neutralized_cells"], 4)
        self.ops.add_formula(path, "D2", "=SUM(A2:A3)", sheet_name="Data")

        with mock.patch.object(
            self.client, "_file_locks", wraps=self.client._file_locks
        ) as file_locks:
            exported = self.ops.csv_export(path, csv_output, sheet_name="Data")

        self.assertTrue(exported["success"])
        self.assertGreaterEqual(exported["neutralized_cells"], 1)
        file_locks.assert_called_once_with(path, csv_output)

        wb = load_workbook(path, data_only=False)
        ws = wb["Data"]
        self.assertEqual(ws["A2"].value, "00123")
        self.assertEqual(ws["B2"].value, "'=1+1")
        self.assertEqual(ws["C2"].value, "'@cmd")
        self.assertEqual(ws["B3"].value, "'+cmd")
        self.assertEqual(ws["C3"].value, "'-cmd")
        wb.close()

        with open(csv_output, newline="", encoding="utf-8") as handle:
            rows = list(csv.reader(handle))
        self.assertEqual(rows[1][1], "'=1+1")
        self.assertEqual(rows[1][3], "'=SUM(A2:A3)")

    def test_xlsx_ops_csv_export_rejects_source_workbook_path(self):
        path = self._path("same_path.xlsx")
        result = self.ops.write_spreadsheet(
            path,
            ["A"],
            [[1]],
            sheet_name="Data",
            overwrite_workbook=True,
        )
        self.assertTrue(result["success"])

        exported = self.ops.csv_export(path, path, sheet_name="Data")

        self.assertFalse(exported["success"])
        self.assertEqual(exported["error_code"], "source_output_same_path")
        wb = load_workbook(path)
        self.assertEqual(wb["Data"]["A2"].value, 1)
        wb.close()

    def test_xlsx_ops_missing_requested_sheet_fails_for_search_and_formula_audit(self):
        path = self._path("missing_sheet.xlsx")
        self.ops.create_spreadsheet(path, sheet_name="Data")

        search = self.ops.search_spreadsheet(path, "anything", sheet_name="Missing")
        audit = self.ops.audit_spreadsheet_formulas(path, sheet_name="Missing")

        self.assertFalse(search["success"])
        self.assertIn("Sheet 'Missing' not found", search["error"])
        self.assertFalse(audit["success"])
        self.assertIn("Sheet 'Missing' not found", audit["error"])

    def test_xlsx_ops_preview_uses_host_spreadsheet_pdf_pipeline(self):
        path = self._path("preview.xlsx")
        output_dir = self._path("preview-out")
        Workbook().save(path)
        captured = {}

        def fake_spreadsheet_to_pdf(file_path, output_path=None):
            self.assertEqual(file_path, path)
            self.assertIsNotNone(output_path)
            with open(output_path, "wb") as handle:
                handle.write(b"%PDF-1.4 fake")
            captured["pdf_path"] = output_path
            return {
                "success": True,
                "input_file": file_path,
                "output_file": output_path,
                "pages": 1,
            }

        def fake_pdf_page_to_image(file_path, render_dir, pages=None, dpi=150, fmt="png"):
            self.assertEqual(file_path, captured["pdf_path"])
            self.assertEqual(render_dir, output_dir)
            self.assertEqual(pages, "0")
            self.assertEqual(dpi, 200)
            self.assertEqual(fmt, "jpg")
            return {
                "success": True,
                "total_pages": 1,
                "pages_rendered": 1,
                "images": [{"page": 0, "file": os.path.join(render_dir, "page_000.jpg")}],
            }

        with mock.patch.object(
            self.client, "spreadsheet_to_pdf", side_effect=fake_spreadsheet_to_pdf
        ):
            with mock.patch.object(
                self.client, "pdf_page_to_image", side_effect=fake_pdf_page_to_image
            ):
                result = self.ops.preview_spreadsheet(
                    path, output_dir, pages="0", dpi=200, fmt="jpg"
                )

        self.assertTrue(result["success"])
        self.assertEqual(result["pages_rendered"], 1)
        self.assertEqual(result["format"], "jpg")
        self.assertFalse(os.path.exists(captured["pdf_path"]))
