import concurrent.futures
import os
import tempfile
import unittest

from openpyxl import load_workbook

from cli_anything.onlyoffice.utils.docserver import get_client


class OnlyOfficeConcurrencyStressTests(unittest.TestCase):
    def setUp(self):
        self.client = get_client()
        self.tmpdir = tempfile.TemporaryDirectory(prefix="oo-stress-")
        self.base = self.tmpdir.name

    def tearDown(self):
        self.tmpdir.cleanup()

    def _path(self, name: str) -> str:
        return os.path.join(self.base, name)

    def test_parallel_spreadsheet_appends_are_consistent(self):
        path = self._path("concurrent_marks.xlsx")
        init = self.client.write_spreadsheet(
            output_path=path,
            headers=["Worker", "Value"],
            data=[],
            sheet_name="Sheet1",
            overwrite_workbook=True,
        )
        self.assertTrue(init["success"])

        operations = 60

        def worker(i: int):
            return self.client.append_to_spreadsheet(
                file_path=path,
                row_data=[f"w{i}", str(i)],
                sheet_name="Sheet1",
            )

        with concurrent.futures.ThreadPoolExecutor(max_workers=12) as ex:
            results = list(ex.map(worker, range(operations)))

        failures = [r for r in results if not r.get("success")]
        self.assertEqual(failures, [])

        wb = load_workbook(path)
        ws = wb["Sheet1"]
        # header row + appended rows
        self.assertEqual(ws.max_row, operations + 1)

    def test_parallel_document_appends_preserve_content(self):
        path = self._path("concurrent.docx")
        created = self.client.create_document(path, "Concurrent Doc", "Start")
        self.assertTrue(created["success"])

        operations = 40

        def worker(i: int):
            return self.client.append_to_document(path, f"line-{i}")

        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as ex:
            results = list(ex.map(worker, range(operations)))

        failures = [r for r in results if not r.get("success")]
        self.assertEqual(failures, [])

        doc = self.client.read_document(path)
        self.assertTrue(doc["success"])
        text = doc["full_text"]
        for i in range(operations):
            self.assertIn(f"line-{i}", text)


if __name__ == "__main__":
    unittest.main()
