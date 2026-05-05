#!/usr/bin/env python3
"""XLSX and chart operations for the OnlyOffice CLI."""

from __future__ import annotations

import os
import re
import tempfile
from pathlib import Path
from typing import Any, Dict, List

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
    from openpyxl.chart.label import DataLabelList
    import openpyxl.utils
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from scipy import stats as scipy_stats
    SCIPY_AVAILABLE = True
except ImportError:
    SCIPY_AVAILABLE = False


class XLSXOperations:
    """Encapsulate spreadsheet, chart, validation, and analysis workflows."""

    def __init__(self, host: Any):
        self.host = host

    def _should_neutralize_spreadsheet_text(self, value: Any) -> bool:
        """Return True for text that spreadsheet apps may treat as a formula."""
        if not isinstance(value, str) or not value:
            return False
        if value.startswith("'"):
            return False
        if value[0] in {"\t", "\r"}:
            return True
        stripped = value.lstrip()
        if not stripped:
            return False
        return stripped[0] in {"=", "+", "-", "@"}

    def _neutralize_spreadsheet_text(self, value: str):
        if self._should_neutralize_spreadsheet_text(value):
            return "'" + value, True
        return value, False

    def _coerce_simple_number(self, value: Any) -> Any:
        if not isinstance(value, str):
            return value
        try:
            if "." in value:
                return float(value)
            return int(value)
        except (ValueError, TypeError):
            return value

    def _prepare_spreadsheet_value(self, value: Any, force_text: bool = False):
        if force_text:
            return self._neutralize_spreadsheet_text("" if value is None else str(value))
        coerced = self._coerce_simple_number(value)
        if isinstance(coerced, str):
            return self._neutralize_spreadsheet_text(coerced)
        return coerced, False

    def _set_prepared_cell(self, cell, value: Any, force_text: bool = False) -> bool:
        prepared, neutralized = self._prepare_spreadsheet_value(value, force_text=force_text)
        cell.value = prepared
        if force_text or neutralized:
            cell.data_type = "s"
        return neutralized

    def _resolve_text_column_indices(
        self, headers: List[Any], text_columns: List[str] = None
    ):
        """Resolve --text-columns entries by case-insensitive header or Excel letter."""
        if not text_columns:
            return set()
        indices = set()
        header_lookup = {}
        for idx, header in enumerate(headers, 1):
            key = str(header).strip().lower()
            if key and key not in header_lookup:
                header_lookup[key] = idx
        for raw_column in text_columns:
            spec = str(raw_column).strip()
            if not spec:
                continue
            header_idx = header_lookup.get(spec.lower())
            if header_idx is not None:
                indices.add(header_idx)
                continue
            if re.fullmatch(r"[A-Za-z]{1,3}", spec):
                try:
                    idx = openpyxl.utils.column_index_from_string(spec.upper())
                except ValueError:
                    continue
                if 1 <= idx <= len(headers):
                    indices.add(idx)
        return indices


    def audit_spreadsheet_formulas(
            self,
            file_path: str,
            sheet_name: str = None,
            max_examples: int = 30,
        ) -> Dict[str, Any]:
            """Audit workbook formulas for unsupported patterns and risk signals."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                wb = load_workbook(file_path, data_only=False)
                if sheet_name and sheet_name not in wb.sheetnames:
                    available = list(wb.sheetnames)
                    wb.close()
                    return {
                        "success": False,
                        "error": f"Sheet '{sheet_name}' not found. Available: {available}",
                    }
                sheets = [sheet_name] if sheet_name else list(wb.sheetnames)

                has_vba = getattr(wb, "vba_archive", None) is not None
                external_link_count = len(getattr(wb, "_external_links", []) or [])

                formula_count = 0
                unsupported_functions = {}
                complex_formulas = []
                external_ref_formulas = []
                function_usage = {}

                for sn in sheets:
                    ws = wb[sn]
                    for row in ws.iter_rows(
                        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
                    ):
                        for cell in row:
                            val = cell.value
                            if not (isinstance(val, str) and val.startswith("=")):
                                continue
                            formula_count += 1
                            funcs = self.host._extract_formula_functions(val)
                            for f in funcs:
                                function_usage[f] = function_usage.get(f, 0) + 1
                                if f not in self.host.supported_formula_functions:
                                    unsupported_functions[f] = (
                                        unsupported_functions.get(f, 0) + 1
                                    )
                            depth = self.host._formula_depth(val)
                            if depth > 3 and len(complex_formulas) < max_examples:
                                complex_formulas.append(
                                    {
                                        "sheet": sn,
                                        "cell": cell.coordinate,
                                        "depth": depth,
                                        "formula": val[:220],
                                    }
                                )
                            if (
                                "[" in val
                                or "http://" in val.lower()
                                or "https://" in val.lower()
                            ) and len(external_ref_formulas) < max_examples:
                                external_ref_formulas.append(
                                    {
                                        "sheet": sn,
                                        "cell": cell.coordinate,
                                        "formula": val[:220],
                                    }
                                )

                wb.close()

                risks = []
                if has_vba:
                    risks.append(
                        "Workbook contains VBA/macros; CLI does not execute macros"
                    )
                if external_link_count > 0 or external_ref_formulas:
                    risks.append(
                        "Workbook contains external links/references that may not resolve in CLI"
                    )
                if unsupported_functions:
                    risks.append(
                        "Workbook uses formula functions outside CLI evaluator support"
                    )
                if complex_formulas:
                    risks.append(
                        "Workbook contains deeply nested formulas that can be error-prone"
                    )

                safe = not risks
                return {
                    "success": True,
                    "file": file_path,
                    "sheet_scope": sheets,
                    "formula_count": formula_count,
                    "has_vba": has_vba,
                    "external_link_count": external_link_count,
                    "function_usage": function_usage,
                    "unsupported_functions": unsupported_functions,
                    "complex_formula_examples": complex_formulas,
                    "external_reference_examples": external_ref_formulas,
                    "safe_for_cli_formula_eval": safe,
                    "risk_level": "low"
                    if safe
                    else ("high" if has_vba or unsupported_functions else "medium"),
                    "risks": risks,
                    "supported_functions": sorted(self.host.supported_formula_functions),
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def create_spreadsheet(
            self, output_path: str, sheet_name: str = "Sheet1"
        ) -> Dict[str, Any]:
            """Create a new .xlsx spreadsheet"""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                with self.host._file_lock(output_path):
                    backup = self.host._snapshot_backup(output_path)
                    wb = Workbook()
                    ws = wb.active
                    ws.title = sheet_name
                    ws.page_setup.paperSize = 9  # A4
                    self.host._safe_save(wb, output_path)
                return {
                    "success": True,
                    "file": output_path,
                    "sheets": [sheet_name],
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def write_spreadsheet(
            self,
            output_path: str,
            headers: List[str],
            data: List[List[Any]],
            sheet_name: str = "Sheet1",
            overwrite_workbook: bool = False,
            coerce_rows: bool = False,
            text_columns: List[str] = None,
        ) -> Dict[str, Any]:
            """Write headers/data to spreadsheet. Non-destructive by default; full overwrite only when explicitly requested.
            text_columns: list of header names whose values should NOT be coerced to numbers (preserves leading zeros)."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                ok, err, data = self.host._validate_tabular_rows(
                    headers, data, coerce_rows=coerce_rows
                )
                if not ok:
                    return {"success": False, "error": err}

                with self.host._file_lock(output_path):
                    backup = self.host._snapshot_backup(output_path)
                    if Path(output_path).exists() and not overwrite_workbook:
                        wb = load_workbook(output_path)
                        ws = (
                            wb[sheet_name]
                            if sheet_name in wb.sheetnames
                            else wb.create_sheet(sheet_name)
                        )
                        ws.delete_rows(1, ws.max_row)
                    else:
                        wb = Workbook()
                        ws = wb.active
                        ws.title = sheet_name

                    neutralized_cells = 0
                    for col, header in enumerate(headers, 1):
                        if self._set_prepared_cell(
                            ws.cell(row=1, column=col),
                            "" if header is None else str(header),
                            force_text=True,
                        ):
                            neutralized_cells += 1
                    text_col_indices = self._resolve_text_column_indices(
                        headers, text_columns
                    )

                    for row_idx, row_data in enumerate(data, 2):
                        for col_idx, value in enumerate(row_data, 1):
                            if self._set_prepared_cell(
                                ws.cell(row=row_idx, column=col_idx),
                                value,
                                force_text=col_idx in text_col_indices,
                            ):
                                neutralized_cells += 1
                    # Auto-fit column widths based on content (min 12, max 50 chars)
                    for col in ws.columns:
                        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                        col_letter = openpyxl.utils.get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 50))
                    # A4 paper size for printing
                    ws.page_setup.paperSize = 9  # 9 = A4
                    self.host._safe_save(wb, output_path)
                return {
                    "success": True,
                    "file": output_path,
                    "rows_written": len(data),
                    "columns": len(headers),
                    "sheet": sheet_name,
                    "mode": "overwrite_workbook" if overwrite_workbook else "update_sheet",
                    "coerce_rows": bool(coerce_rows),
                    "text_columns": sorted(text_col_indices),
                    "neutralized_cells": neutralized_cells,
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def read_spreadsheet(
            self, file_path: str, sheet_name: str = None
        ) -> Dict[str, Any]:
            """Read all data from a spreadsheet"""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                wb = load_workbook(file_path, read_only=True)
                if sheet_name and sheet_name not in wb.sheetnames:
                    wb.close()
                    return {
                        "success": False,
                        "error": f"Sheet '{sheet_name}' not found. Available: {list(wb.sheetnames)}",
                    }
                sheets_to_read = [sheet_name] if sheet_name else list(wb.sheetnames)
                result = {
                    "success": True,
                    "file": file_path,
                    "sheets": wb.sheetnames,
                    "data": {},
                }
                for sn in sheets_to_read:
                    ws = wb[sn]
                    rows_data = []
                    headers = None
                    for row in ws.iter_rows(values_only=True):
                        if all(cell is None for cell in row):
                            continue
                        if headers is None:
                            headers = [str(cell) if cell else "" for cell in row]
                        else:
                            rows_data.append([str(cell) if cell else "" for cell in row])
                    result["data"][sn] = {
                        "headers": headers,
                        "rows": rows_data,
                        "row_count": len(rows_data),
                    }
                wb.close()
                return result
            except Exception as e:
                return {"success": False, "error": str(e)}


    def calculate_column(
            self,
            file_path: str,
            column_letter: str,
            operation: str,
            sheet_name: str = "Sheet1",
            include_formulas: bool = False,
            strict_formula_safety: bool = False,
        ) -> Dict[str, Any]:
            """Calculate column statistics (sum, avg, min, max)"""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                # Must use read_only=False for column access
                wb = load_workbook(file_path)
                ws = self.host._get_sheet(wb, sheet_name)

                # Get column index from letter
                col_idx = openpyxl.utils.column_index_from_string(column_letter.upper())

                # Extract numeric values from the column (skip header row)
                values = []
                formula_rows_total = 0
                formula_rows_evaluated = 0
                formula_rows_failed = 0
                formula_failure_examples = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if col_idx <= len(row):
                        val = row[col_idx - 1]
                        if val is not None and isinstance(val, (int, float)):
                            values.append(val)
                        elif (
                            include_formulas
                            and isinstance(val, str)
                            and val.startswith("=")
                        ):
                            formula_rows_total += 1
                            funcs = self.host._extract_formula_functions(val)
                            unsupported = sorted(
                                [
                                    f
                                    for f in funcs
                                    if f not in self.host.supported_formula_functions
                                ]
                            )
                            if unsupported:
                                formula_rows_failed += 1
                                if len(formula_failure_examples) < 8:
                                    formula_failure_examples.append(
                                        {
                                            "formula": val[:220],
                                            "reason": f"unsupported_functions:{','.join(unsupported)}",
                                        }
                                    )
                                continue
                            resolved = self.host._resolve_formula_value(ws, val[1:])
                            if isinstance(resolved, (int, float)):
                                values.append(float(resolved))
                                formula_rows_evaluated += 1
                            else:
                                formula_rows_failed += 1
                                if len(formula_failure_examples) < 8:
                                    formula_failure_examples.append(
                                        {
                                            "formula": val[:220],
                                            "reason": "evaluation_failed",
                                        }
                                    )

                wb.close()

                if not values:
                    return {
                        "success": False,
                        "error": f"No numeric values in column {column_letter}",
                    }

                op_name = (operation or "all").lower()
                if op_name not in {"sum", "avg", "average", "min", "max", "all"}:
                    return {
                        "success": False,
                        "error": f"Unsupported operation '{operation}'. Use sum|avg|min|max|all",
                    }

                result = {
                    "success": True,
                    "column": column_letter,
                    "sheet": sheet_name,
                    "count": len(values),
                    "sum": sum(values),
                    "average": sum(values) / len(values),
                    "min": min(values),
                    "max": max(values),
                    "formula_mode": bool(include_formulas),
                    "operation": op_name,
                    "formula_rows_total": formula_rows_total,
                    "formula_rows_evaluated": formula_rows_evaluated,
                    "formula_rows_failed": formula_rows_failed,
                    "formula_failure_examples": formula_failure_examples,
                }
                if op_name == "sum":
                    result["value"] = result["sum"]
                elif op_name in {"avg", "average"}:
                    result["value"] = result["average"]
                elif op_name == "min":
                    result["value"] = result["min"]
                elif op_name == "max":
                    result["value"] = result["max"]

                if include_formulas:
                    formula_eval_rate = (
                        (formula_rows_evaluated / formula_rows_total)
                        if formula_rows_total > 0
                        else 1.0
                    )
                    result["formula_eval_rate"] = formula_eval_rate
                    result["formula_reliability"] = (
                        "high"
                        if formula_eval_rate >= 0.95
                        else ("medium" if formula_eval_rate >= 0.7 else "low")
                    )
                    if strict_formula_safety and formula_rows_failed > 0:
                        return {
                            "success": False,
                            "error": "Strict formula safety failed: unresolved/unsupported formulas present",
                            "details": {
                                "formula_rows_total": formula_rows_total,
                                "formula_rows_evaluated": formula_rows_evaluated,
                                "formula_rows_failed": formula_rows_failed,
                                "examples": formula_failure_examples,
                            },
                        }
                return result
            except Exception as e:
                return {"success": False, "error": str(e)}


    def frequencies(
            self,
            file_path: str,
            column_letter: str,
            sheet_name: str = "Sheet1",
            allowed_values: List[str] = None,
        ) -> Dict[str, Any]:
            """Frequency table for one column (counts + percentages)."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                wb = load_workbook(file_path, read_only=True)
                ws = self.host._get_sheet(wb, sheet_name)
                col_idx = openpyxl.utils.column_index_from_string(column_letter.upper())

                counts = {}
                missing = 0
                excluded = 0

                def _allowed(cat):
                    if not allowed_values:
                        return True
                    return any(self.host._value_matches_group(cat, t) for t in allowed_values)

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if col_idx > len(row):
                        missing += 1
                        continue
                    cat = self.host._category_value(row[col_idx - 1])
                    if cat is None:
                        missing += 1
                        continue
                    if not _allowed(cat):
                        excluded += 1
                        continue
                    counts[cat] = counts.get(cat, 0) + 1

                wb.close()

                total_valid = sum(counts.values())
                total_rows = total_valid + missing + excluded
                freq_rows = []
                for k in sorted(counts.keys(), key=lambda x: (str(type(x)), str(x))):
                    c = counts[k]
                    freq_rows.append(
                        {
                            "category": k,
                            "count": c,
                            "percent_valid": (c / total_valid * 100)
                            if total_valid
                            else 0.0,
                            "percent_total": (c / total_rows * 100) if total_rows else 0.0,
                        }
                    )

                return {
                    "success": True,
                    "file": file_path,
                    "sheet": sheet_name,
                    "column": column_letter.upper(),
                    "valid_n": total_valid,
                    "missing_n": missing,
                    "excluded_n": excluded,
                    "total_n": total_rows,
                    "allowed_values": allowed_values or [],
                    "frequencies": freq_rows,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def correlation_test(
            self,
            file_path: str,
            x_column: str,
            y_column: str,
            sheet_name: str = "Sheet1",
            method: str = "pearson",
        ) -> Dict[str, Any]:
            """Correlation between two numeric columns."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            if not SCIPY_AVAILABLE:
                return {"success": False, "error": "scipy not installed"}
            try:
                wb = load_workbook(file_path, read_only=True)
                ws = self.host._get_sheet(wb, sheet_name)
                x_idx = openpyxl.utils.column_index_from_string(x_column.upper())
                y_idx = openpyxl.utils.column_index_from_string(y_column.upper())

                x_vals, y_vals = [], []
                dropped = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    xv = row[x_idx - 1] if x_idx <= len(row) else None
                    yv = row[y_idx - 1] if y_idx <= len(row) else None
                    x_num = self.host._cell_to_float(xv)
                    y_num = self.host._cell_to_float(yv)
                    if x_num is None or y_num is None:
                        dropped += 1
                        continue
                    x_vals.append(x_num)
                    y_vals.append(y_num)
                wb.close()

                if len(x_vals) < 3:
                    return {
                        "success": False,
                        "error": "Need at least 3 paired numeric observations",
                    }

                m = (method or "pearson").lower()
                if m == "pearson":
                    stat, p = scipy_stats.pearsonr(x_vals, y_vals)
                elif m == "spearman":
                    stat, p = scipy_stats.spearmanr(x_vals, y_vals)
                else:
                    return {
                        "success": False,
                        "error": "Unsupported method. Use pearson|spearman",
                    }

                abs_r = abs(float(stat))
                strength = (
                    "negligible"
                    if abs_r < 0.1
                    else (
                        "small" if abs_r < 0.3 else ("moderate" if abs_r < 0.5 else "large")
                    )
                )
                significant = float(p) < 0.05
                direction = (
                    "positive"
                    if float(stat) > 0
                    else ("negative" if float(stat) < 0 else "none")
                )

                return {
                    "success": True,
                    "file": file_path,
                    "sheet": sheet_name,
                    "method": m,
                    "x_column": x_column.upper(),
                    "y_column": y_column.upper(),
                    "n": len(x_vals),
                    "dropped_rows": dropped,
                    "statistic": float(stat),
                    "p_value": float(p),
                    "interpretation": {
                        "alpha": 0.05,
                        "significant": significant,
                        "direction": direction,
                        "strength": strength,
                    },
                    "apa": f"{m.title()} correlation between {x_column.upper()} and {y_column.upper()} was {'ρ' if m == 'spearman' else 'r'} = {float(stat):.3f}, p = {float(p):.4g}, n = {len(x_vals)}.",
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def ttest_independent(
            self,
            file_path: str,
            value_column: str,
            group_column: str,
            group_a: str,
            group_b: str,
            sheet_name: str = "Sheet1",
            equal_var: bool = False,
        ) -> Dict[str, Any]:
            """Independent samples t-test (Welch default)."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            if not SCIPY_AVAILABLE:
                return {"success": False, "error": "scipy not installed"}
            try:
                wb = load_workbook(file_path, read_only=True)
                ws = self.host._get_sheet(wb, sheet_name)
                v_idx = openpyxl.utils.column_index_from_string(value_column.upper())
                g_idx = openpyxl.utils.column_index_from_string(group_column.upper())

                group_a_vals = []
                group_b_vals = []
                dropped = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    gv = row[g_idx - 1] if g_idx <= len(row) else None
                    vv = row[v_idx - 1] if v_idx <= len(row) else None
                    v_num = self.host._cell_to_float(vv)
                    if v_num is None:
                        dropped += 1
                        continue
                    if self.host._value_matches_group(gv, group_a):
                        group_a_vals.append(v_num)
                    elif self.host._value_matches_group(gv, group_b):
                        group_b_vals.append(v_num)
                wb.close()

                if len(group_a_vals) < 2 or len(group_b_vals) < 2:
                    return {
                        "success": False,
                        "error": "Need at least 2 numeric observations in each group",
                    }

                t_stat, p_val = scipy_stats.ttest_ind(
                    group_a_vals,
                    group_b_vals,
                    equal_var=bool(equal_var),
                    nan_policy="omit",
                )

                mean_a = float(sum(group_a_vals) / len(group_a_vals))
                mean_b = float(sum(group_b_vals) / len(group_b_vals))
                sd_a = float(scipy_stats.tstd(group_a_vals))
                sd_b = float(scipy_stats.tstd(group_b_vals))

                # Degrees of freedom
                n_a, n_b = len(group_a_vals), len(group_b_vals)
                if equal_var:
                    df = n_a + n_b - 2
                else:
                    # Welch-Satterthwaite df
                    va, vb = sd_a**2 / n_a, sd_b**2 / n_b
                    df = (va + vb)**2 / (va**2 / (n_a - 1) + vb**2 / (n_b - 1)) if (va + vb) > 0 else n_a + n_b - 2

                # Cohen's d (pooled SD)
                pooled_var = (
                    ((n_a - 1) * (sd_a**2))
                    + ((n_b - 1) * (sd_b**2))
                ) / (n_a + n_b - 2)
                pooled_sd = pooled_var**0.5 if pooled_var > 0 else 0.0
                cohens_d = (mean_a - mean_b) / pooled_sd if pooled_sd else 0.0

                abs_d = abs(float(cohens_d))
                d_mag = (
                    "negligible"
                    if abs_d < 0.2
                    else (
                        "small" if abs_d < 0.5 else ("medium" if abs_d < 0.8 else "large")
                    )
                )
                significant = float(p_val) < 0.05

                return {
                    "success": True,
                    "file": file_path,
                    "sheet": sheet_name,
                    "value_column": value_column.upper(),
                    "group_column": group_column.upper(),
                    "group_a": str(group_a),
                    "group_b": str(group_b),
                    "n_a": n_a,
                    "n_b": n_b,
                    "mean_a": mean_a,
                    "mean_b": mean_b,
                    "sd_a": sd_a,
                    "sd_b": sd_b,
                    "df": round(float(df), 2),
                    "difference": mean_a - mean_b,
                    "equal_var": bool(equal_var),
                    "statistic": float(t_stat),
                    "p_value": float(p_val),
                    "cohens_d": float(cohens_d),
                    "dropped_rows": dropped,
                    "interpretation": {
                        "alpha": 0.05,
                        "significant": significant,
                        "effect_size_magnitude": d_mag,
                        "higher_group": str(group_a)
                        if mean_a > mean_b
                        else (str(group_b) if mean_b > mean_a else "equal"),
                    },
                    "apa": f"{'Welch' if not equal_var else 'Student'}'s independent-samples t-test on {value_column.upper()} by {group_column.upper()} ({group_a} vs {group_b}) found t({df:.1f}) = {float(t_stat):.3f}, p = {float(p_val):.4g}, d = {float(cohens_d):.3f}; M{group_a} = {mean_a:.3f}, M{group_b} = {mean_b:.3f}.",
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def mann_whitney_test(
            self,
            file_path: str,
            value_column: str,
            group_column: str,
            group_a: str,
            group_b: str,
            sheet_name: str = "Sheet1",
        ) -> Dict[str, Any]:
            """Mann-Whitney U test — non-parametric alternative to independent t-test.
            Appropriate for ordinal (Likert) data."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            if not SCIPY_AVAILABLE:
                return {"success": False, "error": "scipy not installed"}
            try:
                wb = load_workbook(file_path, read_only=True)
                ws = self.host._get_sheet(wb, sheet_name)
                val_idx = openpyxl.utils.column_index_from_string(value_column.upper())
                grp_idx = openpyxl.utils.column_index_from_string(group_column.upper())
                group_a_vals, group_b_vals = [], []
                dropped = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    val = row[val_idx - 1] if val_idx - 1 < len(row) else None
                    grp = row[grp_idx - 1] if grp_idx - 1 < len(row) else None
                    if val is None or grp is None:
                        dropped += 1
                        continue
                    try:
                        v = float(val)
                    except (ValueError, TypeError):
                        dropped += 1
                        continue
                    g = str(grp).strip()
                    if g == str(group_a).strip():
                        group_a_vals.append(v)
                    elif g == str(group_b).strip():
                        group_b_vals.append(v)
                wb.close()
                if len(group_a_vals) < 2 or len(group_b_vals) < 2:
                    return {"success": False, "error": f"Insufficient data: group {group_a} n={len(group_a_vals)}, group {group_b} n={len(group_b_vals)}. Need >= 2 per group."}
                u_stat, p_val = scipy_stats.mannwhitneyu(
                    group_a_vals, group_b_vals, alternative="two-sided"
                )
                # Rank-biserial r as effect size
                n1, n2 = len(group_a_vals), len(group_b_vals)
                r_rb = 1 - (2 * float(u_stat)) / (n1 * n2)
                abs_r = abs(r_rb)
                r_mag = "negligible" if abs_r < 0.1 else ("small" if abs_r < 0.3 else ("medium" if abs_r < 0.5 else "large"))
                significant = float(p_val) < 0.05
                import numpy as np
                median_a = float(np.median(group_a_vals))
                median_b = float(np.median(group_b_vals))
                return {
                    "success": True,
                    "file": file_path,
                    "sheet": sheet_name,
                    "test": "Mann-Whitney U",
                    "value_column": value_column.upper(),
                    "group_column": group_column.upper(),
                    "group_a": str(group_a),
                    "group_b": str(group_b),
                    "n_a": n1,
                    "n_b": n2,
                    "median_a": median_a,
                    "median_b": median_b,
                    "statistic": float(u_stat),
                    "p_value": float(p_val),
                    "rank_biserial_r": float(r_rb),
                    "dropped_rows": dropped,
                    "interpretation": {
                        "alpha": 0.05,
                        "significant": significant,
                        "effect_size_magnitude": r_mag,
                        "higher_group": str(group_a) if median_a > median_b else (str(group_b) if median_b > median_a else "equal"),
                    },
                    "apa": f"Mann-Whitney U test on {value_column.upper()} by {group_column.upper()} ({group_a} vs {group_b}) found U = {float(u_stat):.1f}, p = {float(p_val):.4g}, r = {float(r_rb):.3f}; Mdn{group_a} = {median_a:.1f}, Mdn{group_b} = {median_b:.1f}.",
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def chi_square_test(
            self,
            file_path: str,
            row_column: str,
            col_column: str,
            sheet_name: str = "Sheet1",
            row_allowed_values: List[str] = None,
            col_allowed_values: List[str] = None,
        ) -> Dict[str, Any]:
            """Chi-square test of independence for two categorical columns."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            if not SCIPY_AVAILABLE:
                return {"success": False, "error": "scipy not installed"}
            try:
                wb = load_workbook(file_path, read_only=True)
                ws = self.host._get_sheet(wb, sheet_name)
                r_idx = openpyxl.utils.column_index_from_string(row_column.upper())
                c_idx = openpyxl.utils.column_index_from_string(col_column.upper())

                pairs = []
                excluded = 0

                def _allowed(cat, allowed):
                    if not allowed:
                        return True
                    return any(self.host._value_matches_group(cat, t) for t in allowed)

                for row in ws.iter_rows(min_row=2, values_only=True):
                    rv = row[r_idx - 1] if r_idx <= len(row) else None
                    cv = row[c_idx - 1] if c_idx <= len(row) else None
                    r_cat = self.host._category_value(rv)
                    c_cat = self.host._category_value(cv)
                    if r_cat is None or c_cat is None:
                        continue
                    if not _allowed(r_cat, row_allowed_values) or not _allowed(
                        c_cat, col_allowed_values
                    ):
                        excluded += 1
                        continue
                    pairs.append((r_cat, c_cat))
                wb.close()

                if len(pairs) < 2:
                    return {"success": False, "error": "Not enough valid categorical pairs"}

                row_levels = sorted(
                    {p[0] for p in pairs}, key=lambda x: (str(type(x)), str(x))
                )
                col_levels = sorted(
                    {p[1] for p in pairs}, key=lambda x: (str(type(x)), str(x))
                )

                counts = {r: {c: 0 for c in col_levels} for r in row_levels}
                for r, c in pairs:
                    counts[r][c] += 1

                observed = [[counts[r][c] for c in col_levels] for r in row_levels]
                chi2, p, dof, expected = scipy_stats.chi2_contingency(observed)

                n = len(pairs)
                min_dim = min(len(row_levels) - 1, len(col_levels) - 1)
                if min_dim == 0:
                    return {
                        "success": False,
                        "error": f"Cramer's V undefined: one variable has only 1 level (rows={len(row_levels)}, cols={len(col_levels)}). Check --row-valid / --col-valid filters.",
                    }
                cramers_v = ((chi2 / (n * min_dim)) ** 0.5)
                abs_v = abs(float(cramers_v))
                v_mag = (
                    "negligible"
                    if abs_v < 0.1
                    else (
                        "small" if abs_v < 0.3 else ("medium" if abs_v < 0.5 else "large")
                    )
                )
                significant = float(p) < 0.05

                return {
                    "success": True,
                    "file": file_path,
                    "sheet": sheet_name,
                    "row_column": row_column.upper(),
                    "col_column": col_column.upper(),
                    "n": n,
                    "excluded_n": excluded,
                    "row_allowed_values": row_allowed_values or [],
                    "col_allowed_values": col_allowed_values or [],
                    "rows": row_levels,
                    "cols": col_levels,
                    "observed": observed,
                    "expected": [list(map(float, r)) for r in expected.tolist()],
                    "degrees_of_freedom": int(dof),
                    "statistic": float(chi2),
                    "p_value": float(p),
                    "cramers_v": float(cramers_v),
                    "interpretation": {
                        "alpha": 0.05,
                        "significant": significant,
                        "association_strength": v_mag,
                    },
                    "apa": f"Chi-square test of {row_column.upper()} by {col_column.upper()} was χ²({int(dof)}) = {float(chi2):.3f}, p = {float(p):.4g}, Cramer's V = {float(cramers_v):.3f}, n = {n}.",
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def open_text_extract(
            self,
            file_path: str,
            column_letter: str,
            sheet_name: str = "Sheet1",
            limit: int = 20,
            min_length: int = 20,
        ) -> Dict[str, Any]:
            """Extract non-empty open-text responses from a column for qualitative analysis."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                wb = load_workbook(file_path, read_only=True)
                ws = self.host._get_sheet(wb, sheet_name)
                col_idx = openpyxl.utils.column_index_from_string(column_letter.upper())

                responses = []
                all_non_empty = 0
                for r in range(2, (ws.max_row or 0) + 1):
                    val = ws.cell(r, col_idx).value
                    if val is None:
                        continue
                    text = str(val).strip()
                    if not text:
                        continue
                    all_non_empty += 1
                    if len(text) < int(min_length):
                        continue
                    responses.append(
                        {
                            "row": r,
                            "cell": f"{column_letter.upper()}{r}",
                            "text": text,
                            "length": len(text),
                        }
                    )

                wb.close()
                responses = responses[: max(1, int(limit))]
                return {
                    "success": True,
                    "file": file_path,
                    "sheet": sheet_name,
                    "column": column_letter.upper(),
                    "total_non_empty": all_non_empty,
                    "returned": len(responses),
                    "limit": int(limit),
                    "min_length": int(min_length),
                    "responses": responses,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def open_text_keywords(
            self,
            file_path: str,
            column_letter: str,
            sheet_name: str = "Sheet1",
            top_n: int = 15,
            min_word_length: int = 4,
        ) -> Dict[str, Any]:
            """Get keyword frequency summary from an open-text column for theme seeding."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                wb = load_workbook(file_path, read_only=True)
                ws = self.host._get_sheet(wb, sheet_name)
                col_idx = openpyxl.utils.column_index_from_string(column_letter.upper())

                stopwords = {
                    "that",
                    "this",
                    "with",
                    "have",
                    "from",
                    "more",
                    "would",
                    "their",
                    "they",
                    "what",
                    "before",
                    "into",
                    "about",
                    "being",
                    "because",
                    "through",
                    "there",
                    "where",
                    "when",
                    "which",
                    "just",
                    "also",
                    "will",
                    "than",
                    "then",
                    "your",
                    "them",
                    "some",
                    "very",
                    "much",
                    "need",
                    "skills",
                    "health",
                    "experience",
                    "placement",
                    "clinical",
                    "students",
                    "student",
                    "graduate",
                    "university",
                    "degree",
                    "work",
                    "working",
                    "field",
                    "practical",
                    "able",
                    "feel",
                    "think",
                    "would",
                    "could",
                    "like",
                    "more",
                    "been",
                    "have",
                    "being",
                    "really",
                    "well",
                }

                token_counts = {}
                response_count = 0
                for r in range(2, (ws.max_row or 0) + 1):
                    val = ws.cell(r, col_idx).value
                    if val is None:
                        continue
                    text = str(val).strip().lower()
                    if not text:
                        continue
                    response_count += 1
                    tokens = re.findall(r"[a-zA-Z][a-zA-Z\-']+", text)
                    for t in tokens:
                        if len(t) < int(min_word_length):
                            continue
                        if t in stopwords:
                            continue
                        token_counts[t] = token_counts.get(t, 0) + 1
                wb.close()

                top = sorted(token_counts.items(), key=lambda kv: kv[1], reverse=True)[
                    : max(1, int(top_n))
                ]
                return {
                    "success": True,
                    "file": file_path,
                    "sheet": sheet_name,
                    "column": column_letter.upper(),
                    "response_count": response_count,
                    "top_n": int(top_n),
                    "min_word_length": int(min_word_length),
                    "keywords": [
                        {
                            "keyword": k,
                            "count": c,
                            "percent_responses": (c / response_count * 100)
                            if response_count
                            else 0.0,
                        }
                        for k, c in top
                    ],
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def research_analysis_pack(
            self,
            file_path: str,
            sheet_name: str = "Sheet0",
            profile: str = "hlth3112",
            require_formula_safe: bool = False,
        ) -> Dict[str, Any]:
            """Run a standardized spreadsheet analysis bundle for research assignments."""
            profile_key = (profile or "hlth3112").lower()
            if profile_key != "hlth3112":
                return {
                    "success": False,
                    "error": "Unsupported profile. Use: hlth3112",
                }

            result = {
                "success": True,
                "file": file_path,
                "sheet": sheet_name,
                "profile": profile_key,
                "steps": {
                    "formula_audit": [],
                    "frequencies": [],
                    "descriptives": [],
                    "correlations": [],
                    "ttests": [],
                    "chi2": [],
                    "qualitative": [],
                },
                "summary": {},
            }

            # Profile configuration tuned for HLTH3112/HLTH3011 survey structure.
            freq_specs = [
                {"column": "A", "valid": ["1", "2"], "label": "Gender"},
                {
                    "column": "F",
                    "valid": ["1", "2", "3", "4", "5"],
                    "label": "Professional_placement",
                },
                {
                    "column": "AL",
                    "valid": ["1", "2", "3", "4", "5"],
                    "label": "More_experience",
                },
            ]
            descriptive_specs = [
                {"column": "M", "op": "all", "label": "Confident_job"},
                {"column": "R", "op": "all", "label": "Overall_prepared"},
                {"column": "AL", "op": "all", "label": "More_experience"},
                {"column": "AM", "op": "all", "label": "More_training"},
            ]
            corr_specs = [
                {
                    "x": "M",
                    "y": "R",
                    "method": "spearman",
                    "label": "Confident_job_vs_Overall_prepared",
                },
                {
                    "x": "Y",
                    "y": "M",
                    "method": "spearman",
                    "label": "Interpret_statistics_vs_Confident_job",
                },
            ]
            # Mann-Whitney U (non-parametric) for ordinal Likert data
            mannwhitney_specs = [
                {
                    "value": "M",
                    "group": "F",
                    "a": "1",
                    "b": "4",
                    "label": "Confident_job_by_placement_1_vs_4",
                },
                {
                    "value": "R",
                    "group": "F",
                    "a": "1",
                    "b": "4",
                    "label": "Overall_prepared_by_placement_1_vs_4",
                },
            ]
            # Parametric t-test (kept for reference/comparison — agent can run separately)
            ttest_specs = []
            chi_specs = [
                {
                    "row": "A",
                    "col": "AL",
                    "row_valid": ["1", "2"],
                    "col_valid": ["1", "2", "3", "4", "5"],
                    "label": "Gender_vs_More_experience",
                }
            ]
            qualitative_specs = [
                {"column": "AN", "label": "Strongest_skill"},
                {"column": "AO", "label": "Underprepared_skill"},
                {"column": "AP", "label": "Employers_value"},
                {"column": "AQ", "label": "Improve_employability"},
            ]

            def _capture(bucket: str, payload: Dict[str, Any], meta: Dict[str, Any]):
                entry = {"meta": meta, "result": payload}
                result["steps"][bucket].append(entry)

            audit = self.audit_spreadsheet_formulas(
                file_path=file_path,
                sheet_name=sheet_name,
                max_examples=20,
            )
            _capture("formula_audit", audit, {"sheet": sheet_name})

            if require_formula_safe:
                if not audit.get("success"):
                    return {
                        "success": False,
                        "file": file_path,
                        "sheet": sheet_name,
                        "profile": profile_key,
                        "steps": result["steps"],
                        "summary": {
                            "total_analyses": 1,
                            "succeeded": 0,
                            "failed": 1,
                            "completion_rate": 0.0,
                            "formula_audit_safe": None,
                        },
                        "error": "Formula safety audit failed and strict policy is enabled",
                    }
                if not audit.get("safe_for_cli_formula_eval", False):
                    return {
                        "success": False,
                        "file": file_path,
                        "sheet": sheet_name,
                        "profile": profile_key,
                        "steps": result["steps"],
                        "summary": {
                            "total_analyses": 1,
                            "succeeded": 0,
                            "failed": 1,
                            "completion_rate": 0.0,
                            "formula_audit_safe": False,
                        },
                        "error": "Formula safety policy blocked execution: workbook is not safe_for_cli_formula_eval",
                        "formula_audit": audit,
                    }

            for spec in freq_specs:
                payload = self.frequencies(
                    file_path=file_path,
                    column_letter=spec["column"],
                    sheet_name=sheet_name,
                    allowed_values=spec.get("valid"),
                )
                _capture("frequencies", payload, spec)

            for spec in descriptive_specs:
                payload = self.calculate_column(
                    file_path=file_path,
                    column_letter=spec["column"],
                    operation=spec.get("op", "all"),
                    sheet_name=sheet_name,
                    include_formulas=False,
                )
                _capture("descriptives", payload, spec)

            for spec in corr_specs:
                payload = self.correlation_test(
                    file_path=file_path,
                    x_column=spec["x"],
                    y_column=spec["y"],
                    sheet_name=sheet_name,
                    method=spec.get("method", "pearson"),
                )
                _capture("correlations", payload, spec)

            for spec in ttest_specs:
                payload = self.ttest_independent(
                    file_path=file_path,
                    value_column=spec["value"],
                    group_column=spec["group"],
                    group_a=spec["a"],
                    group_b=spec["b"],
                    sheet_name=sheet_name,
                    equal_var=False,
                )
                _capture("ttests", payload, spec)

            # Mann-Whitney U tests (non-parametric, appropriate for ordinal/Likert)
            if "mannwhitney" not in result["steps"]:
                result["steps"]["mannwhitney"] = []
            for spec in mannwhitney_specs:
                payload = self.mann_whitney_test(
                    file_path=file_path,
                    value_column=spec["value"],
                    group_column=spec["group"],
                    group_a=spec["a"],
                    group_b=spec["b"],
                    sheet_name=sheet_name,
                )
                _capture("mannwhitney", payload, spec)

            for spec in chi_specs:
                payload = self.chi_square_test(
                    file_path=file_path,
                    row_column=spec["row"],
                    col_column=spec["col"],
                    sheet_name=sheet_name,
                    row_allowed_values=spec.get("row_valid"),
                    col_allowed_values=spec.get("col_valid"),
                )
                _capture("chi2", payload, spec)

            for spec in qualitative_specs:
                kw = self.open_text_keywords(
                    file_path=file_path,
                    column_letter=spec["column"],
                    sheet_name=sheet_name,
                    top_n=12,
                    min_word_length=4,
                )
                ex = self.open_text_extract(
                    file_path=file_path,
                    column_letter=spec["column"],
                    sheet_name=sheet_name,
                    limit=6,
                    min_length=35,
                )
                payload = {
                    "success": bool(kw.get("success") and ex.get("success")),
                    "keywords": kw,
                    "quotes": ex,
                }
                if not payload["success"]:
                    payload["error"] = kw.get("error") or ex.get("error")
                _capture("qualitative", payload, spec)

            all_results = []
            for bucket in result["steps"].values():
                all_results.extend(bucket)
            ok = sum(1 for x in all_results if x["result"].get("success"))
            fail = len(all_results) - ok

            sig_corr = 0
            for item in result["steps"].get("correlations", []):
                if item["result"].get("interpretation", {}).get("significant"):
                    sig_corr += 1
            sig_t = 0
            for item in result["steps"].get("ttests", []):
                if item["result"].get("interpretation", {}).get("significant"):
                    sig_t += 1
            sig_chi = 0
            for item in result["steps"].get("chi2", []):
                if item["result"].get("interpretation", {}).get("significant"):
                    sig_chi += 1

            result["summary"] = {
                "total_analyses": len(all_results),
                "succeeded": ok,
                "failed": fail,
                "completion_rate": (ok / len(all_results)) if all_results else 0.0,
                "significant": {
                    "correlations": sig_corr,
                    "ttests": sig_t,
                    "chi2": sig_chi,
                },
                "formula_audit_safe": bool(audit.get("safe_for_cli_formula_eval"))
                if audit.get("success")
                else None,
                "require_formula_safe": bool(require_formula_safe),
            }
            if ok == 0:
                result["success"] = False
                result["error"] = (
                    "All analyses failed. Check sheet name and dataset structure."
                )

            return result


    def append_to_spreadsheet(
            self, file_path: str, row_data: List[Any], sheet_name: str = None
        ) -> Dict[str, Any]:
            """Append a row to an existing spreadsheet"""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active

                    # Find the next empty row
                    next_row = ws.max_row + 1

                    neutralized_cells = 0
                    for col_idx, value in enumerate(row_data, 1):
                        if self._set_prepared_cell(
                            ws.cell(row=next_row, column=col_idx), value
                        ):
                            neutralized_cells += 1

                    self.host._safe_save(wb, file_path)
                    sheet_title = ws.title
                    wb.close()
                return {
                    "success": True,
                    "file": file_path,
                    "row_added": next_row,
                    "columns": len(row_data),
                    "sheet": sheet_title,
                    "neutralized_cells": neutralized_cells,
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def search_spreadsheet(
            self, file_path: str, search_text: str, sheet_name: str = None
        ) -> Dict[str, Any]:
            """Search for text in a spreadsheet"""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                wb = load_workbook(file_path, read_only=True)
                if sheet_name and sheet_name not in wb.sheetnames:
                    available = list(wb.sheetnames)
                    wb.close()
                    return {
                        "success": False,
                        "error": f"Sheet '{sheet_name}' not found. Available: {available}",
                    }
                sheets_to_search = [sheet_name] if sheet_name else list(wb.sheetnames)

                results = []
                for sn in sheets_to_search:
                    ws = wb[sn]
                    for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
                        for col_idx, cell in enumerate(row, 1):
                            if (
                                cell.value is not None
                                and search_text.lower() in str(cell.value).lower()
                            ):
                                col_letter = openpyxl.utils.get_column_letter(col_idx)
                                results.append(
                                    {
                                        "sheet": sn,
                                        "cell": f"{col_letter}{row_idx}",
                                        "value": str(cell.value),
                                        "row": row_idx,
                                        "column": col_letter,
                                    }
                                )

                wb.close()
                return {
                    "success": True,
                    "file": file_path,
                    "search_text": search_text,
                    "matches": len(results),
                    "results": results,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def add_formula(
            self, file_path: str, cell: str, formula: str, sheet_name: str = "Sheet1"
        ) -> Dict[str, Any]:
            """Add formula to a specific cell"""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name)
                    ws[cell] = formula
                    self.host._safe_save(wb, file_path)
                    wb.close()
                return {
                    "success": True,
                    "file": file_path,
                    "cell": cell,
                    "formula": formula,
                    "sheet": sheet_name,
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def create_chart(
            self,
            file_path: str,
            chart_type: str,
            data_range: str,
            categories_range: str = None,
            title: str = "Chart",
            sheet_name: str = None,
            output_sheet: str = None,
            x_label: str = None,
            y_label: str = None,
            show_data_labels: bool = False,
            show_legend: bool = True,
            legend_pos: str = "right",
            colors: List[str] = None,
        ) -> Dict[str, Any]:
            """
            Create a chart in a spreadsheet.

            Args:
                file_path: Path to the Excel file
                chart_type: Type of chart ('bar', 'line', 'pie', 'scatter')
                data_range: Cell range for data values (e.g., 'B2:D5')
                categories_range: Cell range for category labels (e.g., 'A2:A5')
                title: Chart title
                sheet_name: Source sheet name (default: active sheet)
                output_sheet: Sheet to place chart (default: same as source)
                x_label: X-axis label
                y_label: Y-axis label
                show_data_labels: Show data labels on chart
                show_legend: Show legend
                legend_pos: Legend position ('right', 'top', 'bottom', 'left')
                colors: List of hex colors for chart series

            Returns:
                Dict with success status and chart details
            """
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}

            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active

                    # Validate ranges before creating chart
                    _ = self.host._parse_range(data_range)
                    if categories_range:
                        _ = self.host._parse_range(categories_range)
                    if not self.host._range_has_data(ws, data_range):
                        return {
                            "success": False,
                            "error": f"Data range {data_range} contains no usable values",
                        }
                    if categories_range and not self.host._range_has_data(ws, categories_range):
                        return {
                            "success": False,
                            "error": f"Category range {categories_range} contains no usable values",
                        }

                    # Create the appropriate chart type
                    chart = self.host._create_chart_object(chart_type)

                    # Parse ranges and create references
                    if categories_range:
                        min_col, min_row, max_col, max_row = self.host._parse_range(
                            categories_range
                        )
                        cat_ref = Reference(
                            ws,
                            min_col=min_col,
                            min_row=min_row,
                            max_col=max_col,
                            max_row=max_row,
                        )
                    else:
                        # Default: use first column of data range
                        dmin_col, dmin_row, dmax_col, dmax_row = self.host._parse_range(
                            data_range
                        )
                        cat_ref = Reference(
                            ws,
                            min_col=dmin_col,
                            min_row=dmin_row + 1,
                            max_col=dmin_col,
                            max_row=dmax_row,
                        )

                    # Data values (Y-axis)
                    min_col, min_row, max_col, max_row = self.host._parse_range(data_range)
                    data_ref = Reference(
                        ws,
                        min_col=min_col,
                        min_row=min_row,
                        max_col=max_col,
                        max_row=max_row,
                    )

                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cat_ref)

                    chart.title = title

                    if not show_legend:
                        chart.has_legend = False
                    else:
                        chart.legend.pos = legend_pos

                    if x_label:
                        chart.x_axis.title = x_label
                    if y_label:
                        chart.y_axis.title = y_label

                    if show_data_labels:
                        chart.dataLabels = DataLabelList()
                        chart.dataLabels.show_val = True

                    if colors:
                        self.host._apply_chart_colors(chart, colors)

                    if output_sheet and output_sheet != ws.title:
                        if output_sheet not in wb.sheetnames:
                            out_ws = wb.create_sheet(output_sheet)
                        else:
                            out_ws = wb[output_sheet]
                        out_ws.add_chart(chart, "A2")
                    else:
                        ws.add_chart(chart, "A10")

                    self.host._safe_save(wb, file_path)
                    wb.close()

                return {
                    "success": True,
                    "file": file_path,
                    "chart_type": chart_type,
                    "title": title,
                    "sheet": ws.title,
                    "data_range": data_range,
                    "categories_range": categories_range,
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def create_comparison_chart(
            self,
            file_path: str,
            chart_type: str,
            sheet_name: str = None,
            title: str = "Comparison Chart",
            start_row: int = 1,
            start_col: int = 1,
            num_categories: int = None,
            num_series: int = None,
            category_col: int = 1,
            value_cols: List[int] = None,
            output_cell: str = "A10",
            show_data_labels: bool = False,
            show_legend: bool = True,
        ) -> Dict[str, Any]:
            """
            Create a chart from structured data in a spreadsheet.

            Args:
                file_path: Path to the Excel file
                chart_type: Type of chart ('bar', 'line', 'pie', 'scatter')
                sheet_name: Source sheet name
                title: Chart title
                start_row: Starting row of data (1-indexed, usually 2 to skip header)
                start_col: Starting column of data (1-indexed)
                num_categories: Number of category rows (default: auto-detect)
                num_series: Number of data series (columns)
                category_col: Column containing category labels (1-indexed)
                value_cols: List of columns containing values (1-indexed)
                output_cell: Cell where chart should be placed
                show_data_labels: Show data labels
                show_legend: Show legend

            Returns:
                Dict with success status and chart details
            """
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}

            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active

                    # Auto-detect ranges if not specified
                    if num_categories is None:
                        num_categories = ws.max_row - start_row + 1
                    if num_categories <= 0:
                        return {
                            "success": False,
                            "error": "No category rows found for chart",
                        }

                    if value_cols is None:
                        value_cols = list(range(start_col + 1, ws.max_column + 1))
                    if not value_cols:
                        return {
                            "success": False,
                            "error": "No value columns found for chart",
                        }

                    if num_series is None:
                        num_series = len(value_cols)

                    chart = self.host._create_chart_object(chart_type)

                    cat_min_row = start_row
                    cat_max_row = start_row + num_categories - 1
                    cat_ref = Reference(
                        ws,
                        min_col=category_col,
                        min_row=cat_min_row,
                        max_col=category_col,
                        max_row=cat_max_row,
                    )

                    for col in value_cols:
                        data_ref = Reference(
                            ws,
                            min_col=col,
                            min_row=cat_min_row,
                            max_col=col,
                            max_row=cat_max_row,
                        )
                        chart.add_data(data_ref, titles_from_data=True)

                    chart.set_categories(cat_ref)
                    chart.title = title
                    chart.has_legend = show_legend

                    if show_data_labels:
                        chart.dataLabels = DataLabelList()
                        chart.dataLabels.show_val = True

                    ws.add_chart(chart, output_cell)
                    self.host._safe_save(wb, file_path)
                    wb.close()

                return {
                    "success": True,
                    "file": file_path,
                    "chart_type": chart_type,
                    "title": title,
                    "sheet": ws.title,
                    "categories": num_categories,
                    "series": num_series,
                    "output_cell": output_cell,
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def create_grade_distribution_chart(
            self,
            file_path: str,
            sheet_name: str = None,
            grade_column: str = "B",
            title: str = "Grade Distribution",
            output_cell: str = "F2",
        ) -> Dict[str, Any]:
            """
            Create a pie chart showing grade distribution.

            Args:
                file_path: Path to the Excel file
                sheet_name: Sheet name
                grade_column: Column letter containing grades
                title: Chart title
                output_cell: Cell where chart should be placed

            Returns:
                Dict with success status and chart details
            """
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}

            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active

                    col_idx = openpyxl.utils.column_index_from_string(grade_column)
                    grade_counts = {}
                    total_rows = 0

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if col_idx <= len(row) and row[col_idx - 1]:
                            grade = str(row[col_idx - 1]).strip().upper()
                            grade_counts[grade] = grade_counts.get(grade, 0) + 1
                            total_rows += 1

                    if not grade_counts:
                        return {
                            "success": False,
                            "error": "No grades found in column " + grade_column,
                        }

                    chart = PieChart()
                    chart.title = title
                    # Write chart helper data to a separate sheet to avoid corrupting the data sheet
                    helper_name = "_chart_data"
                    if helper_name not in wb.sheetnames:
                        helper_ws = wb.create_sheet(helper_name)
                    else:
                        helper_ws = wb[helper_name]
                    # Find next free row in helper sheet
                    base_row = (helper_ws.max_row or 0) + 2
                    helper_ws.cell(row=base_row, column=1, value="Grade")
                    helper_ws.cell(row=base_row, column=2, value="Count")
                    for i, (grade, count) in enumerate(sorted(grade_counts.items()), 1):
                        helper_ws.cell(row=base_row + i, column=1, value=grade)
                        helper_ws.cell(row=base_row + i, column=2, value=count)

                    cat_ref = Reference(
                        helper_ws,
                        min_col=1,
                        min_row=base_row + 1,
                        max_col=1,
                        max_row=base_row + len(grade_counts),
                    )
                    data_ref = Reference(
                        helper_ws,
                        min_col=2,
                        min_row=base_row + 1,
                        max_col=2,
                        max_row=base_row + len(grade_counts),
                    )

                    chart.add_data(data_ref, titles_from_data=False)
                    chart.set_categories(cat_ref)
                    chart.dataLabels = DataLabelList()
                    chart.dataLabels.show_val = True
                    chart.dataLabels.show_pct = True

                    ws.add_chart(chart, output_cell)
                    self.host._safe_save(wb, file_path)
                    wb.close()

                return {
                    "success": True,
                    "file": file_path,
                    "chart_type": "pie",
                    "title": title,
                    "sheet": ws.title,
                    "total_grades": total_rows,
                    "distribution": grade_counts,
                    "output_cell": output_cell,
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def create_progress_chart(
            self,
            file_path: str,
            student_column: str = "A",
            grade_column: str = "B",
            sheet_name: str = None,
            title: str = "Student Progress",
            output_cell: str = "D2",
            show_data_labels: bool = True,
        ) -> Dict[str, Any]:
            """
            Create a bar chart showing individual student grades.

            Args:
                file_path: Path to the Excel file
                student_column: Column letter with student names
                grade_column: Column letter with grades
                sheet_name: Sheet name
                title: Chart title
                output_cell: Cell where chart should be placed
                show_data_labels: Show data labels on bars

            Returns:
                Dict with success status and chart details
            """
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}

            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active

                    student_col_idx = openpyxl.utils.column_index_from_string(
                        student_column
                    )
                    grade_col_idx = openpyxl.utils.column_index_from_string(grade_column)
                    num_rows = ws.max_row - 1

                    if num_rows <= 0:
                        return {"success": False, "error": "No data found"}

                    chart = BarChart()
                    chart.type = "bar"
                    chart.title = title

                    cat_ref = Reference(
                        ws,
                        min_col=student_col_idx,
                        min_row=2,
                        max_col=student_col_idx,
                        max_row=num_rows + 1,
                    )

                    data_ref = Reference(
                        ws,
                        min_col=grade_col_idx,
                        min_row=2,
                        max_col=grade_col_idx,
                        max_row=num_rows + 1,
                    )

                    chart.add_data(data_ref, titles_from_data=False)
                    chart.set_categories(cat_ref)
                    chart.x_axis.title = "Grade"
                    chart.y_axis.title = "Student"

                    if show_data_labels:
                        chart.dataLabels = DataLabelList()
                        chart.dataLabels.show_val = True

                    ws.add_chart(chart, output_cell)
                    self.host._safe_save(wb, file_path)
                    wb.close()

                return {
                    "success": True,
                    "file": file_path,
                    "chart_type": "bar_horizontal",
                    "title": title,
                    "sheet": ws.title,
                    "students": num_rows,
                    "output_cell": output_cell,
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def cell_read(
            self, file_path: str, cell_ref: str, sheet_name: str = None
        ) -> Dict[str, Any]:
            """Read a single cell value by reference (e.g., 'B5')."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active
                cell = ws[cell_ref]
                value = cell.value
                wb.close()
                return {
                    "success": True, "file": file_path,
                    "sheet": ws.title, "cell": cell_ref,
                    "value": value, "type": type(value).__name__,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def cell_write(
            self, file_path: str, cell_ref: str, value: str,
            sheet_name: str = None, as_text: bool = False,
        ) -> Dict[str, Any]:
            """Write a value to a single cell."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active
                    cell = ws[cell_ref]
                    neutralized = self._set_prepared_cell(
                        cell, value, force_text=as_text
                    )
                    stored_value = cell.value
                    self.host._safe_save(wb, file_path)
                    sheet_title = ws.title
                    wb.close()
                return {
                    "success": True, "file": file_path,
                    "sheet": sheet_title, "cell": cell_ref,
                    "value": stored_value,
                    "neutralized": neutralized,
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def range_read(
            self, file_path: str, range_ref: str, sheet_name: str = None
        ) -> Dict[str, Any]:
            """Read a range of cells (e.g., 'A1:D10')."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active
                min_col, min_row, max_col, max_row = self.host._parse_range(range_ref)
                rows = []
                for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                         min_col=min_col, max_col=max_col, values_only=True):
                    rows.append([str(c) if c is not None else "" for c in row])
                wb.close()
                return {
                    "success": True, "file": file_path,
                    "sheet": ws.title, "range": range_ref,
                    "rows": rows, "row_count": len(rows),
                    "col_count": max_col - min_col + 1,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def delete_rows(
            self, file_path: str, start_row: int, count: int = 1, sheet_name: str = None
        ) -> Dict[str, Any]:
            """Delete rows from spreadsheet. start_row is 1-indexed."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active
                    ws.delete_rows(start_row, count)
                    self.host._safe_save(wb, file_path)
                    wb.close()
                return {
                    "success": True, "file": file_path,
                    "sheet": ws.title, "deleted_from_row": start_row,
                    "rows_deleted": count, "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def delete_columns(
            self, file_path: str, start_col: int, count: int = 1, sheet_name: str = None
        ) -> Dict[str, Any]:
            """Delete columns from spreadsheet. start_col is 1-indexed."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active
                    ws.delete_cols(start_col, count)
                    self.host._safe_save(wb, file_path)
                    wb.close()
                return {
                    "success": True, "file": file_path,
                    "sheet": ws.title, "deleted_from_col": start_col,
                    "cols_deleted": count, "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def sort_sheet(
            self, file_path: str, column_letter: str, sheet_name: str = None,
            descending: bool = False, numeric: bool = False,
        ) -> Dict[str, Any]:
            """Sort sheet data by a column. Preserves header row."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active
                    col_idx = openpyxl.utils.column_index_from_string(column_letter.upper()) - 1
                    # Read all data
                    all_rows = list(ws.iter_rows(values_only=True))
                    if len(all_rows) < 2:
                        wb.close()
                        return {"success": False, "error": "Not enough data to sort (need header + data)"}
                    header = all_rows[0]
                    data = all_rows[1:]

                    def sort_key(row):
                        val = row[col_idx] if col_idx < len(row) else None
                        if val is None:
                            return (1, "")
                        if numeric:
                            try:
                                return (0, float(val))
                            except (ValueError, TypeError):
                                return (1, str(val).lower())
                        return (0, str(val).lower())

                    data.sort(key=sort_key, reverse=descending)
                    # Clear and rewrite
                    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                        for cell in row_cells:
                            cell.value = None
                    for ci, h in enumerate(header, 1):
                        ws.cell(row=1, column=ci, value=h)
                    for ri, row in enumerate(data, 2):
                        for ci, val in enumerate(row, 1):
                            ws.cell(row=ri, column=ci, value=val)
                    self.host._safe_save(wb, file_path)
                    wb.close()
                return {
                    "success": True, "file": file_path,
                    "sheet": ws.title, "sorted_by": column_letter.upper(),
                    "descending": descending, "rows_sorted": len(data),
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def filter_rows(
            self, file_path: str, column_letter: str, operator: str, value: str,
            sheet_name: str = None,
        ) -> Dict[str, Any]:
            """Filter/query rows. Returns matching rows without modifying the file.
            operator: eq|ne|gt|lt|ge|le|contains|startswith|endswith"""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            valid_operators = {"eq", "ne", "gt", "lt", "ge", "le", "contains", "startswith", "endswith"}
            if operator not in valid_operators:
                return {
                    "success": False,
                    "error": f"Unknown operator '{operator}'. Valid: {' | '.join(sorted(valid_operators))}",
                }
            try:
                wb = load_workbook(file_path, read_only=True)
                ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active
                col_idx = openpyxl.utils.column_index_from_string(column_letter.upper()) - 1
                all_rows = list(ws.iter_rows(values_only=True))
                wb.close()
                if not all_rows:
                    return {"success": False, "error": "Empty sheet"}
                header = [str(h) if h else "" for h in all_rows[0]]
                data = all_rows[1:]

                def compare(cell_val):
                    if cell_val is None:
                        return False
                    cv = str(cell_val)
                    if operator in ("gt", "lt", "ge", "le"):
                        try:
                            nv = float(cv)
                            tv = float(value)
                            if operator == "gt": return nv > tv
                            if operator == "lt": return nv < tv
                            if operator == "ge": return nv >= tv
                            if operator == "le": return nv <= tv
                        except (ValueError, TypeError):
                            return False
                    if operator == "eq":
                        try:
                            return float(cv) == float(value)
                        except (ValueError, TypeError):
                            return cv.lower() == value.lower()
                    if operator == "ne":
                        try:
                            return float(cv) != float(value)
                        except (ValueError, TypeError):
                            return cv.lower() != value.lower()
                    if operator == "contains": return value.lower() in cv.lower()
                    if operator == "startswith": return cv.lower().startswith(value.lower())
                    if operator == "endswith": return cv.lower().endswith(value.lower())
                    return False

                matched = []
                for ri, row in enumerate(data, 2):
                    cell_val = row[col_idx] if col_idx < len(row) else None
                    if compare(cell_val):
                        matched.append({
                            "row_number": ri,
                            "values": [str(c) if c is not None else "" for c in row],
                        })
                return {
                    "success": True, "file": file_path,
                    "sheet": ws.title, "filter": f"{column_letter.upper()} {operator} {value}",
                    "headers": header, "total_rows": len(data),
                    "matched_rows": len(matched),
                    "rows": matched[:500],
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def sheet_list(self, file_path: str) -> Dict[str, Any]:
            """List all sheets in a workbook with row/column counts."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                wb = load_workbook(file_path, read_only=True)
                sheets = []
                for name in wb.sheetnames:
                    ws = wb[name]
                    sheets.append({
                        "name": name,
                        "max_row": ws.max_row,
                        "max_column": ws.max_column,
                    })
                wb.close()
                return {
                    "success": True, "file": file_path,
                    "sheet_count": len(sheets), "sheets": sheets,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def sheet_add(
            self, file_path: str, sheet_name: str, position: int = None
        ) -> Dict[str, Any]:
            """Add a new sheet to the workbook."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    if sheet_name in wb.sheetnames:
                        wb.close()
                        return {"success": False, "error": f"Sheet '{sheet_name}' already exists"}
                    if position is not None:
                        wb.create_sheet(sheet_name, position)
                    else:
                        wb.create_sheet(sheet_name)
                    self.host._safe_save(wb, file_path)
                    wb.close()
                return {
                    "success": True, "file": file_path,
                    "sheet_added": sheet_name,
                    "sheets": wb.sheetnames if hasattr(wb, 'sheetnames') else [sheet_name],
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def sheet_delete(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
            """Delete a sheet from the workbook."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    if sheet_name not in wb.sheetnames:
                        wb.close()
                        return {"success": False, "error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"}
                    if len(wb.sheetnames) == 1:
                        wb.close()
                        return {"success": False, "error": "Cannot delete the only sheet in the workbook"}
                    del wb[sheet_name]
                    self.host._safe_save(wb, file_path)
                    remaining = wb.sheetnames
                    wb.close()
                return {
                    "success": True, "file": file_path,
                    "sheet_deleted": sheet_name,
                    "remaining_sheets": remaining,
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def sheet_rename(
            self, file_path: str, old_name: str, new_name: str
        ) -> Dict[str, Any]:
            """Rename a sheet."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    if old_name not in wb.sheetnames:
                        wb.close()
                        return {"success": False, "error": f"Sheet '{old_name}' not found. Available: {wb.sheetnames}"}
                    wb[old_name].title = new_name
                    self.host._safe_save(wb, file_path)
                    wb.close()
                return {
                    "success": True, "file": file_path,
                    "old_name": old_name, "new_name": new_name,
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def merge_cells(
            self, file_path: str, range_ref: str, sheet_name: str = None
        ) -> Dict[str, Any]:
            """Merge a range of cells."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active
                    ws.merge_cells(range_ref)
                    self.host._safe_save(wb, file_path)
                    wb.close()
                return {
                    "success": True, "file": file_path,
                    "sheet": ws.title, "merged_range": range_ref,
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def unmerge_cells(
            self, file_path: str, range_ref: str, sheet_name: str = None
        ) -> Dict[str, Any]:
            """Unmerge a range of cells."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active
                    ws.unmerge_cells(range_ref)
                    self.host._safe_save(wb, file_path)
                    wb.close()
                return {
                    "success": True, "file": file_path,
                    "sheet": ws.title, "unmerged_range": range_ref,
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def format_cells(
            self, file_path: str, range_ref: str, sheet_name: str = None,
            bold: bool = False, italic: bool = False, font_name: str = None,
            font_size: int = None, color: str = None, bg_color: str = None,
            number_format: str = None, alignment: str = None, wrap_text: bool = False,
        ) -> Dict[str, Any]:
            """Apply formatting to a range of cells."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                from openpyxl.styles import Font, PatternFill, Alignment, numbers
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active
                    min_col, min_row, max_col, max_row = self.host._parse_range(range_ref)
                    count = 0
                    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                             min_col=min_col, max_col=max_col):
                        for cell in row:
                            if bold or italic or font_name or font_size or color:
                                f = cell.font.copy(
                                    bold=bold if bold else cell.font.bold,
                                    italic=italic if italic else cell.font.italic,
                                    name=font_name if font_name else cell.font.name,
                                    size=font_size if font_size else cell.font.size,
                                    color=color.lstrip("#") if color else cell.font.color,
                                )
                                cell.font = f
                            if bg_color:
                                cell.fill = PatternFill(
                                    start_color=bg_color.lstrip("#"),
                                    end_color=bg_color.lstrip("#"),
                                    fill_type="solid",
                                )
                            if number_format:
                                cell.number_format = number_format
                            if alignment or wrap_text:
                                h_align = alignment if alignment else None
                                cell.alignment = Alignment(
                                    horizontal=h_align,
                                    wrap_text=wrap_text,
                                )
                            count += 1
                    self.host._safe_save(wb, file_path)
                    wb.close()
                return {
                    "success": True, "file": file_path,
                    "sheet": ws.title, "range": range_ref,
                    "cells_formatted": count, "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def csv_import(
            self, file_path: str, csv_path: str, sheet_name: str = "Sheet1",
            delimiter: str = ",", text_columns: List[str] = None,
        ) -> Dict[str, Any]:
            """Import a CSV file into an xlsx spreadsheet."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                import csv as csv_module
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    if Path(file_path).exists():
                        wb = load_workbook(file_path)
                        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
                        ws.delete_rows(1, ws.max_row)
                    else:
                        wb = Workbook()
                        ws = wb.active
                        ws.title = sheet_name
                    with open(csv_path, "r", newline="", encoding="utf-8-sig") as f:
                        reader = csv_module.reader(f, delimiter=delimiter)
                        row_count = 0
                        neutralized_cells = 0
                        text_col_indices = set()
                        for ri, row in enumerate(reader, 1):
                            if ri == 1:
                                text_col_indices = self._resolve_text_column_indices(
                                    row, text_columns
                                )
                            for ci, val in enumerate(row, 1):
                                if self._set_prepared_cell(
                                    ws.cell(row=ri, column=ci),
                                    val,
                                    force_text=ci in text_col_indices,
                                ):
                                    neutralized_cells += 1
                            row_count += 1
                    self.host._safe_save(wb, file_path)
                    wb.close()
                return {
                    "success": True, "file": file_path,
                    "csv_source": csv_path, "sheet": sheet_name,
                    "rows_imported": row_count,
                    "text_columns": sorted(text_col_indices),
                    "neutralized_cells": neutralized_cells,
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def csv_export(
            self, file_path: str, csv_path: str, sheet_name: str = None,
            delimiter: str = ",",
        ) -> Dict[str, Any]:
            """Export a spreadsheet sheet to CSV."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                import csv as csv_module
                target = Path(csv_path)
                if Path(file_path).resolve() == target.resolve():
                    return {
                        "success": False,
                        "error": "CSV export output path must be different from the source workbook.",
                        "error_code": "source_output_same_path",
                        "file": file_path,
                        "csv_output": csv_path,
                    }
                target.parent.mkdir(parents=True, exist_ok=True)
                row_count = 0
                neutralized_cells = 0
                with self.host._file_locks(file_path, csv_path):
                    wb = load_workbook(file_path, read_only=True, data_only=False)
                    try:
                        ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active
                        sheet_title = ws.title
                        fd, tmp_path = tempfile.mkstemp(
                            prefix=f".{target.name}.", dir=str(target.parent)
                        )
                        os.close(fd)
                        try:
                            with open(tmp_path, "w", newline="", encoding="utf-8") as f:
                                writer = csv_module.writer(f, delimiter=delimiter)
                                for row in ws.iter_rows(values_only=True):
                                    csv_row = []
                                    for value in row:
                                        if value is None:
                                            csv_row.append("")
                                            continue
                                        if isinstance(value, str):
                                            value, neutralized = self._neutralize_spreadsheet_text(value)
                                            if neutralized:
                                                neutralized_cells += 1
                                            csv_row.append(value)
                                        else:
                                            csv_row.append(str(value))
                                    writer.writerow(csv_row)
                                    row_count += 1
                            with open(tmp_path, "rb") as handle:
                                os.fsync(handle.fileno())
                            os.replace(tmp_path, str(target))
                            self.host._fsync_directory(target.parent)
                        finally:
                            if os.path.exists(tmp_path):
                                os.unlink(tmp_path)
                    finally:
                        wb.close()
                return {
                    "success": True, "file": file_path,
                    "csv_output": csv_path, "sheet": sheet_title,
                    "rows_exported": row_count,
                    "neutralized_cells": neutralized_cells,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def add_validation(
            self, file_path: str, cell_range: str,
            validation_type: str, operator: str = None,
            formula1: str = None, formula2: str = None,
            allow_blank: bool = True, sheet_name: str = None,
            error_message: str = None, error_title: str = None,
            prompt_message: str = None, prompt_title: str = None,
            error_style: str = "stop",
        ) -> Dict[str, Any]:
            """Add a data validation rule to a cell range.
            Types: list, whole, decimal, date, time, textLength, custom.
            Operators: between, notBetween, equal, notEqual, lessThan,
                       lessThanOrEqual, greaterThan, greaterThanOrEqual.
            For list type: formula1 is comma-separated values (e.g. 'Yes,No,Maybe').
            error_style: stop (reject), warning (warn+allow), information (info+allow)."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            valid_types = {"list", "whole", "decimal", "date", "time", "textLength", "custom"}
            if validation_type not in valid_types:
                return {"success": False, "error": f"Unknown type '{validation_type}'. Valid: {sorted(valid_types)}"}
            valid_ops = {"between", "notBetween", "equal", "notEqual", "lessThan",
                         "lessThanOrEqual", "greaterThan", "greaterThanOrEqual"}
            if operator and operator not in valid_ops:
                return {"success": False, "error": f"Unknown operator '{operator}'. Valid: {sorted(valid_ops)}"}
            try:
                from openpyxl.worksheet.datavalidation import DataValidation
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active

                    # Build formula1 for list type
                    f1 = formula1
                    if validation_type == "list" and f1 and not f1.startswith('"') and not f1.startswith("="):
                        f1 = f'"{f1}"'

                    dv = DataValidation(
                        type=validation_type,
                        operator=operator,
                        formula1=f1,
                        formula2=formula2,
                        allow_blank=allow_blank,
                    )
                    if error_message:
                        dv.error = error_message
                        dv.showErrorMessage = True
                    if error_title:
                        dv.errorTitle = error_title
                        dv.showErrorMessage = True
                    if prompt_message:
                        dv.prompt = prompt_message
                        dv.showInputMessage = True
                    if prompt_title:
                        dv.promptTitle = prompt_title
                        dv.showInputMessage = True
                    if error_style:
                        dv.errorStyle = error_style

                    dv.add(cell_range)
                    ws.add_data_validation(dv)
                    self.host._safe_save(wb, file_path)
                return {
                    "success": True, "file": file_path,
                    "sheet": ws.title, "range": cell_range,
                    "type": validation_type, "operator": operator,
                    "formula1": f1, "formula2": formula2,
                    "error_style": error_style,
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def add_dropdown(
            self, file_path: str, cell_range: str, options: str,
            allow_blank: bool = True, sheet_name: str = None,
            prompt: str = None, error_message: str = None,
        ) -> Dict[str, Any]:
            """Shortcut: add a dropdown list validation. Options are comma-separated."""
            return self.add_validation(
                file_path, cell_range, validation_type="list",
                formula1=options, allow_blank=allow_blank,
                sheet_name=sheet_name,
                prompt_message=prompt, prompt_title="Select",
                error_message=error_message or f"Must be one of: {options}",
                error_title="Invalid Selection",
            )


    def list_validations(
            self, file_path: str, sheet_name: str = None,
        ) -> Dict[str, Any]:
            """List all data validation rules on a sheet."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                wb = load_workbook(file_path)
                ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active
                rules = []
                for dv in ws.data_validations.dataValidation:
                    rule = {
                        "range": str(dv.sqref),
                        "type": dv.type,
                        "operator": dv.operator,
                        "formula1": dv.formula1,
                        "formula2": dv.formula2,
                        "allow_blank": dv.allowBlank,
                        "error_style": dv.errorStyle,
                        "error_message": dv.error,
                        "error_title": dv.errorTitle,
                        "prompt": dv.prompt,
                        "prompt_title": dv.promptTitle,
                    }
                    # For list type, parse the allowed values
                    if dv.type == "list" and dv.formula1:
                        raw = dv.formula1.strip('"')
                        rule["allowed_values"] = [v.strip() for v in raw.split(",")]
                    rules.append(rule)
                wb.close()
                return {
                    "success": True, "file": file_path,
                    "sheet": ws.title,
                    "validation_count": len(rules),
                    "validations": rules,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def remove_validation(
            self, file_path: str, cell_range: str = None,
            sheet_name: str = None, remove_all: bool = False,
        ) -> Dict[str, Any]:
            """Remove data validation from a range or all validations on a sheet."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            if not cell_range and not remove_all:
                return {"success": False, "error": "Provide --range <range> or --all to remove validations"}
            try:
                with self.host._file_lock(file_path):
                    backup = self.host._snapshot_backup(file_path)
                    wb = load_workbook(file_path)
                    ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active
                    removed = 0
                    if remove_all:
                        removed = len(ws.data_validations.dataValidation)
                        ws.data_validations.dataValidation.clear()
                    else:
                        target = cell_range.upper()
                        to_keep = []
                        for dv in ws.data_validations.dataValidation:
                            if str(dv.sqref).upper() == target:
                                removed += 1
                            else:
                                to_keep.append(dv)
                        ws.data_validations.dataValidation.clear()
                        for dv in to_keep:
                            ws.add_data_validation(dv)
                    self.host._safe_save(wb, file_path)
                return {
                    "success": True, "file": file_path,
                    "sheet": ws.title,
                    "removed": removed,
                    "remaining": len(ws.data_validations.dataValidation),
                    "backup": backup or None,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def validate_data(
            self, file_path: str, sheet_name: str = None,
            max_rows: int = 1000,
        ) -> Dict[str, Any]:
            """Audit existing cell data against validation rules. Returns pass/fail per cell."""
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxl not installed"}
            try:
                wb = load_workbook(file_path)
                ws = self.host._get_sheet(wb, sheet_name) if sheet_name else wb.active
                results = []
                total_checked = 0
                total_pass = 0
                total_fail = 0
                for dv in ws.data_validations.dataValidation:
                    for cell_range_str in str(dv.sqref).split():
                        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(cell_range_str)
                        max_row = min(max_row, min_row + max_rows - 1)
                        for row in range(min_row, max_row + 1):
                            for col in range(min_col, max_col + 1):
                                cell = ws.cell(row=row, column=col)
                                if cell.value is None:
                                    if not dv.allowBlank:
                                        total_checked += 1
                                        total_fail += 1
                                        results.append({
                                            "cell": cell.coordinate,
                                            "value": None,
                                            "valid": False,
                                            "reason": "Blank not allowed",
                                            "rule_type": dv.type,
                                        })
                                    continue
                                total_checked += 1
                                valid, reason = self.host._check_validation(cell.value, dv)
                                if valid:
                                    total_pass += 1
                                else:
                                    total_fail += 1
                                    results.append({
                                        "cell": cell.coordinate,
                                        "value": str(cell.value)[:100],
                                        "valid": False,
                                        "reason": reason,
                                        "rule_type": dv.type,
                                        "rule_range": str(dv.sqref),
                                    })
                wb.close()
                return {
                    "success": True, "file": file_path,
                    "sheet": ws.title,
                    "cells_checked": total_checked,
                    "cells_passed": total_pass,
                    "cells_failed": total_fail,
                    "failures": results,
                }
            except Exception as e:
                return {"success": False, "error": str(e)}


    def spreadsheet_to_pdf(
            self, file_path: str, output_path: str = None,
        ) -> Dict[str, Any]:
            """Convert a spreadsheet file to PDF via OnlyOffice x2t inside Docker."""
            return self.host._office_to_pdf(file_path, output_path=output_path)


    def preview_spreadsheet(
            self,
            file_path: str,
            output_dir: str,
            pages: str = None,
            dpi: int = 150,
            fmt: str = "png",
        ) -> Dict[str, Any]:
            """Render spreadsheet pages as images via OnlyOffice conversion + PyMuPDF."""
            return self.host._preview_via_pdf(
                file_path,
                output_dir,
                self.host.spreadsheet_to_pdf,
                pages=pages,
                dpi=dpi,
                fmt=fmt,
            )
